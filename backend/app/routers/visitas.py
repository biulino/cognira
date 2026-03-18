from datetime import datetime
from math import radians, sin, cos, sqrt, atan2
from typing import Optional
from io import BytesIO

from pydantic import BaseModel

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, or_, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.deps import get_current_user, require_role, tenant_filter
from app.models.analyst import Analista
from app.models.establishment import Estabelecimento
from app.models.study import Onda
from app.models.visit import Visita, CampoVisita, CaracterizacaoCache
from app.models.photo import FotoVisita
from app.models.user import Utilizador
from app.schemas import VisitaCreate, VisitaOut, VisitaEstadoUpdate, VisitaListResponse
from app.edition import require_pro
from app.services.state_machine import transition_visita
from app.services import pii
from app.services.audit import log_action
from app.services.visitas_service import (
    compute_visita_stats,
    compute_visita_timeline,
    detect_fraude,
    compute_visita_sla,
)

router = APIRouter()


@router.get("/barcode")
async def lookup_barcode(
    code: str,
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Look up a barcode/EAN code in shelf audit items and visit campos."""
    from app.models.shelf_audit import ShelfAuditItem

    # 1. Search shelf_audit_items by EAN (most authoritative source)
    shelf_rows = (
        await db.execute(
            select(ShelfAuditItem)
            .where(ShelfAuditItem.ean == code)
            .order_by(ShelfAuditItem.criado_em.desc())
            .limit(5)
        )
    ).scalars().all()

    if shelf_rows:
        latest = shelf_rows[0]
        return {
            "code": code,
            "found": True,
            "name": latest.produto_nome,
            "brand": None,
            "description": f"Encontrado em {len(shelf_rows)} auditoria(s). Último preço: {latest.preco_real}€" if latest.preco_real else f"Encontrado em {len(shelf_rows)} auditoria(s).",
        }

    # 2. Fallback: search campos_visita for barcode-tagged entries
    cv_rows = (
        await db.execute(
            select(CampoVisita)
            .where(
                func.lower(CampoVisita.chave).in_(["ean", "barcode", "código", "codigo", "ean13", "upc"]),
                CampoVisita.valor == code,
            )
            .limit(3)
        )
    ).scalars().all()

    if cv_rows:
        return {
            "code": code,
            "found": True,
            "name": None,
            "brand": None,
            "description": f"Código registado em {len(cv_rows)} visita(s).",
        }

    return {"code": code, "found": False}


@router.get("/", response_model=VisitaListResponse)
async def list_visitas(
    estudo_id: Optional[int] = None,
    estado: Optional[str] = None,
    onda_id: Optional[int] = None,
    analista_id: Optional[int] = None,
    estabelecimento_id: Optional[int] = None,
    search: Optional[str] = None,
    data_inicio: Optional[datetime] = None,
    data_fim: Optional[datetime] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=1000),
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    is_staff = user.role_global in ("admin", "coordenador")
    base_q = select(Visita)
    tid = tenant_filter(user)
    if not is_staff:
        # Restrict to estudos the user has an explicit permission for
        allowed_estudo_ids = [p.estudo_id for p in (user.permissoes or [])]
        if not allowed_estudo_ids:
            return {"items": [], "total": 0}
        base_q = base_q.where(Visita.estudo_id.in_(allowed_estudo_ids))
    elif tid is not None:
        # Staff: scope by tenant via Estudo → Cliente chain
        from app.models.study import Estudo
        from app.models.client import Cliente
        base_q = (
            base_q
            .join(Estudo, Visita.estudo_id == Estudo.id)
            .join(Cliente, Estudo.cliente_id == Cliente.id)
            .where(Cliente.tenant_id == tid)
        )
    if estudo_id:
        base_q = base_q.where(Visita.estudo_id == estudo_id)
    if estado:
        base_q = base_q.where(Visita.estado == estado)
    if onda_id:
        base_q = base_q.where(Visita.onda_id == onda_id)
    if analista_id:
        base_q = base_q.where(Visita.analista_id == analista_id)
    if estabelecimento_id:
        base_q = base_q.where(Visita.estabelecimento_id == estabelecimento_id)
    if search:
        try:
            search_id = int(search)
            base_q = base_q.where(Visita.id == search_id)
        except ValueError:
            # Text search: match against establishment name
            base_q = base_q.join(
                Estabelecimento, Visita.estabelecimento_id == Estabelecimento.id
            ).where(Estabelecimento.nome.ilike(f"%{search}%"))
    if data_inicio:
        base_q = base_q.where(
            or_(Visita.planeada_em >= data_inicio, Visita.realizada_inicio >= data_inicio)
        )
    if data_fim:
        base_q = base_q.where(
            or_(Visita.planeada_em <= data_fim, Visita.realizada_inicio <= data_fim)
        )

    # Count total
    count_q = select(func.count()).select_from(base_q.subquery())
    total = (await db.execute(count_q)).scalar_one()

    # Paginate
    q = base_q.order_by(Visita.id.desc()).offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(q)
    visitas = result.scalars().all()

    # Batch pre-fetch related data to avoid N+1 queries
    analista_ids = {v.analista_id for v in visitas if v.analista_id}
    estab_ids = {v.estabelecimento_id for v in visitas}
    onda_ids = {v.onda_id for v in visitas if v.onda_id}
    visita_ids_pg = [v.id for v in visitas]

    analista_map: dict[int, Analista] = {}
    if analista_ids:
        res = await db.execute(select(Analista).where(Analista.id.in_(analista_ids)))
        analista_map = {a.id: a for a in res.scalars().all()}

    estab_map: dict[int, Estabelecimento] = {}
    if estab_ids:
        res = await db.execute(select(Estabelecimento).where(Estabelecimento.id.in_(estab_ids)))
        estab_map = {e.id: e for e in res.scalars().all()}

    onda_map: dict[int, Onda] = {}
    if onda_ids:
        res = await db.execute(select(Onda).where(Onda.id.in_(onda_ids)))
        onda_map = {o.id: o for o in res.scalars().all()}

    cc_map: dict[int, CaracterizacaoCache] = {}
    if visita_ids_pg:
        res = await db.execute(select(CaracterizacaoCache).where(CaracterizacaoCache.visita_id.in_(visita_ids_pg)))
        cc_map = {cc.visita_id: cc for cc in res.scalars().all()}

    foto_count_map: dict[int, int] = {}
    if visita_ids_pg:
        fc_rows = await db.execute(
            select(FotoVisita.visita_id, func.count(FotoVisita.id).label("cnt"))
            .where(FotoVisita.visita_id.in_(visita_ids_pg))
            .group_by(FotoVisita.visita_id)
        )
        foto_count_map = {row.visita_id: row.cnt for row in fc_rows.all()}

    items = []
    for v in visitas:
        item = VisitaOut.model_validate(v)
        a = analista_map.get(v.analista_id) if v.analista_id else None
        if a:
            raw = a.nome
            item.analista_nome = pii.decrypt(raw) if isinstance(raw, (bytes, bytearray)) else str(raw or "")
            item.analista_codigo = a.codigo_externo
        e = estab_map.get(v.estabelecimento_id)
        if e:
            item.estabelecimento_nome = e.nome
        o = onda_map.get(v.onda_id) if v.onda_id else None
        if o:
            item.onda_label = o.label
        cc = cc_map.get(v.id)
        if cc and cc.dados:
            item.caracterizacao = cc.dados
        item.fotos_count = foto_count_map.get(v.id, 0)
        items.append(item)

    return VisitaListResponse(items=items, total=total, page=page, page_size=page_size)


@router.get("/stats")
async def visitas_stats(
    estudo_id: Optional[int] = None,
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await compute_visita_stats(db, user, estudo_id)


@router.get("/timeline")
async def visitas_timeline(
    days: int = Query(30, ge=7, le=365),
    estudo_id: Optional[int] = None,
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await compute_visita_timeline(db, user, days, estudo_id)


# ---------------------------------------------------------------------------
# Fraud / Anomaly Detection — heuristic engine  (MUST be before /{visita_id})
# ---------------------------------------------------------------------------

@router.get("/fraude", dependencies=[Depends(require_role("admin", "coordenador"))])
async def detectar_fraude(
    estudo_id: Optional[int] = None,
    min_intervalo_minutos: int = Query(30, ge=5, le=120, description="Threshold (min) between consecutive visits considered suspicious"),
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await detect_fraude(db, user, estudo_id, min_intervalo_minutos)


# ---------------------------------------------------------------------------
# SLA Monitor — GET /visitas/sla  (MUST be before /{visita_id})
# ---------------------------------------------------------------------------

@router.get("/sla", dependencies=[Depends(require_role("admin", "coordenador"))])
async def visitas_sla(
    estudo_id: Optional[int] = None,
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await compute_visita_sla(db, user, estudo_id)


# ---------------------------------------------------------------------------
# Criterios score breakdown — GET /visitas/criterios-score  (before /{visita_id})
# ---------------------------------------------------------------------------

@router.get("/criterios-score")
async def criterios_score_breakdown(
    estudo_id: int,
    onda_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    user: Utilizador = Depends(get_current_user),
):
    """Per-criterion conformance/score breakdown for drill-down charts."""
    from collections import defaultdict
    from app.models.evaluation import Grelha, CriterioGrelha, RespostaVisita
    from app.models.study import Estudo as _EstudoC
    from app.models.client import Cliente as _ClienteC

    # Verify tenant ownership of the requested estudo
    tid = tenant_filter(user)
    if tid is not None:
        ownership = (await db.execute(
            select(_EstudoC.id)
            .join(_ClienteC, _EstudoC.cliente_id == _ClienteC.id)
            .where(_EstudoC.id == estudo_id, _ClienteC.tenant_id == tid)
        )).scalar_one_or_none()
        if ownership is None:
            raise HTTPException(status_code=403, detail="Acesso negado")

    grelhas = (await db.execute(
        select(Grelha).where(Grelha.estudo_id == estudo_id)
    )).scalars().all()
    if not grelhas:
        return []

    grelha_ids = [g.id for g in grelhas]
    criterios = (await db.execute(
        select(CriterioGrelha).where(CriterioGrelha.grelha_id.in_(grelha_ids))
    )).scalars().all()
    if not criterios:
        return []

    criterio_map = {c.id: c for c in criterios}
    criterio_ids = [c.id for c in criterios]

    vq = select(Visita.id).where(
        Visita.estudo_id == estudo_id,
        Visita.estado.in_(["inserida", "validada", "fechada"]),
    )
    if onda_id:
        vq = vq.where(Visita.onda_id == onda_id)
    visita_ids_result = (await db.execute(vq)).scalars().all()
    if not visita_ids_result:
        return []

    respostas = (await db.execute(
        select(RespostaVisita).where(
            RespostaVisita.visita_id.in_(visita_ids_result),
            RespostaVisita.criterio_id.in_(criterio_ids),
        )
    )).scalars().all()

    by_criterio: dict[int, list] = defaultdict(list)
    for r in respostas:
        by_criterio[r.criterio_id].append(r.valor)

    result_list = []
    for cid, valores in by_criterio.items():
        c = criterio_map.get(cid)
        if not c:
            continue
        total = len(valores)
        if c.tipo == "boolean":
            conformes = sum(1 for v in valores if str(v).lower() in ("true", "1", "sim", "yes"))
            pct: float | None = round(conformes / total * 100, 1) if total else 0.0
            avg: float | None = None
        elif c.tipo == "escala":
            numeric = []
            for v in valores:
                if v is not None:
                    try:
                        numeric.append(float(v))
                    except (ValueError, TypeError):
                        pass
            avg = round(sum(numeric) / len(numeric), 1) if numeric else None
            pct = round(avg * 10, 1) if avg is not None else None
        else:
            avg = None
            pct = None
        result_list.append({
            "criterio_id": cid,
            "label": c.label,
            "tipo": c.tipo,
            "peso": float(c.peso) if c.peso else None,
            "total_respostas": total,
            "conformidade_pct": pct,
            "score_medio": avg,
        })

    result_list.sort(key=lambda x: (x["conformidade_pct"] or 0))
    return result_list


@router.get("/{visita_id}", response_model=VisitaOut)
async def get_visita(
    visita_id: int,
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from app.models.study import Estudo as _EstudoGet, Onda as _OndaGet
    from app.models.client import Cliente as _ClienteGet

    # Single query: fetch visita + all related objects via outerjoin
    stmt = (
        select(Visita, Analista, Estabelecimento, Onda, CaracterizacaoCache)
        .outerjoin(Analista, Analista.id == Visita.analista_id)
        .outerjoin(Estabelecimento, Estabelecimento.id == Visita.estabelecimento_id)
        .outerjoin(Onda, Onda.id == Visita.onda_id)
        .outerjoin(CaracterizacaoCache, CaracterizacaoCache.visita_id == Visita.id)
        .where(Visita.id == visita_id)
    )
    row = (await db.execute(stmt)).one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Visita não encontrada")
    visita, analista, estab, onda, cc = row

    # Tenant isolation: verify ownership via Estudo → Cliente
    tid = tenant_filter(user)
    if tid is not None:
        ownership = (await db.execute(
            select(_ClienteGet.id)
            .join(_EstudoGet, _EstudoGet.cliente_id == _ClienteGet.id)
            .where(_EstudoGet.id == visita.estudo_id, _ClienteGet.tenant_id == tid)
        )).scalar_one_or_none()
        if ownership is None:
            raise HTTPException(status_code=403, detail="Acesso negado")

    item = VisitaOut.model_validate(visita)
    if analista:
        raw = analista.nome
        item.analista_nome = pii.decrypt(raw) if isinstance(raw, (bytes, bytearray)) else str(raw or "")
        item.analista_codigo = analista.codigo_externo
    if estab:
        item.estabelecimento_nome = estab.nome
    if onda:
        item.onda_label = onda.label
    if cc and cc.dados:
        item.caracterizacao = cc.dados
    return item


@router.post("/", response_model=VisitaOut, status_code=201)
async def create_visita(
    body: VisitaCreate,
    user: Utilizador = Depends(require_role("admin", "coordenador")),
    db: AsyncSession = Depends(get_db),
):
    # Plan limit: max_visitas_mes
    if user.tenant and user.tenant.plano and user.tenant.plano.max_visitas_mes is not None:
        from datetime import date, timezone as tz
        from app.models.client import Cliente
        from app.models.study import Estudo
        hoje = date.today()
        month_start = hoje.replace(day=1)
        count = (await db.execute(
            select(func.count(Visita.id))
            .join(Estudo, Estudo.id == Visita.estudo_id)
            .join(Cliente, Cliente.id == Estudo.cliente_id)
            .where(
                Cliente.tenant_id == user.tenant_id,
                Visita.criado_em >= month_start,
            )
        )).scalar_one()
        if count >= user.tenant.plano.max_visitas_mes:
            raise HTTPException(
                status_code=402,
                detail=f"Limite de visitas mensais do plano atingido ({user.tenant.plano.max_visitas_mes}).",
            )
    visita = Visita(**body.model_dump())
    db.add(visita)
    await db.flush()

    # Auto-assign grelha from tipo_visita when not explicitly provided
    if not visita.grelha_id and visita.tipo_visita and visita.tipo_visita not in ("normal", "extra"):
        from app.models.evaluation import Grelha
        grid = (await db.execute(
            select(Grelha).where(
                Grelha.estudo_id == visita.estudo_id,
                Grelha.tipo_visita == visita.tipo_visita,
            )
        )).scalar_one_or_none()
        if grid:
            visita.grelha_id = grid.id
            await db.flush()

    await db.refresh(visita)

    try:
        from app.services.webhooks import dispatch_event
        import asyncio
        asyncio.create_task(dispatch_event(db, "visita.criada", {
            "visita_id": visita.id,
            "estudo_id": visita.estudo_id,
            "estado": visita.estado,
        }))
    except Exception:
        pass

    return visita


@router.put("/{visita_id}/estado", response_model=VisitaOut)
async def update_visita_estado(
    visita_id: int,
    body: VisitaEstadoUpdate,
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Visita).where(Visita.id == visita_id))
    visita = result.scalar_one_or_none()
    if not visita:
        raise HTTPException(status_code=404, detail="Visita não encontrada")
    # Tenant isolation
    _tid_upd = tenant_filter(user)
    if _tid_upd is not None:
        from app.models.study import Estudo as _EstudoU
        from app.models.client import Cliente as _ClienteU
        _own = (await db.execute(
            select(_ClienteU.id)
            .join(_EstudoU, _EstudoU.cliente_id == _ClienteU.id)
            .where(_EstudoU.id == visita.estudo_id, _ClienteU.tenant_id == _tid_upd)
        )).scalar_one_or_none()
        if _own is None:
            raise HTTPException(status_code=403, detail="Acesso negado")

    old_estado = visita.estado
    visita = await transition_visita(visita, body.estado, user, db, body.motivo_anulacao)
    await log_action(
        db,
        utilizador_id=user.id,
        entidade="Visita",
        entidade_id=str(visita_id),
        acao="estado_transition",
        dados_anteriores={"estado": old_estado},
        dados_novos={"estado": body.estado},
    )

    # Webhook dispatch
    try:
        from app.services.webhooks import dispatch_event
        import asyncio
        evento = "visita.concluida" if body.estado in ("validada", "fechada") else "visita.estado_mudou"
        asyncio.create_task(dispatch_event(db, evento, {
            "visita_id": visita.id,
            "estudo_id": visita.estudo_id,
            "old_estado": old_estado,
            "new_estado": body.estado,
        }))
    except Exception:
        pass

    # Notify analista by email on important state changes
    if body.estado in ("corrigir", "corrigir_email", "validada", "fechada") and visita.analista_id:
        try:
            from app.services.email import send_estado_change
            from app.models.analyst import Analista
            from app.models.study import Estudo as EstudoModel
            from app.services.push_service import send_push as _send_push

            analista = (await db.execute(select(Analista).where(Analista.id == visita.analista_id))).scalar_one_or_none()
            estudo = (await db.execute(select(EstudoModel).where(EstudoModel.id == visita.estudo_id))).scalar_one_or_none()

            if analista and analista.email:
                import asyncio
                analista_nome = pii.decrypt(analista.nome) if isinstance(analista.nome, (bytes, bytearray)) else str(analista.nome or "")
                analista_email = pii.decrypt(analista.email) if isinstance(analista.email, (bytes, bytearray)) else str(analista.email or "")
                asyncio.create_task(send_estado_change(
                    to_email=analista_email,
                    visita_id=visita.id,
                    estudo_nome=estudo.nome if estudo else f"#{visita.estudo_id}",
                    analista_nome=analista_nome,
                    old_estado=old_estado,
                    new_estado=body.estado,
                ))

            # Web Push for the analyst's linked user account (if any)
            if analista and analista.utilizador_id:
                import asyncio as _asyncio2
                _push_title = f"Visita #{visita.id} — {body.estado}"
                if body.estado == "validada":
                    _push_body = "A tua visita foi validada com sucesso."
                elif body.estado in ("corrigir", "corrigir_email"):
                    _push_body = "A tua visita foi devolvida para correcção."
                else:
                    _push_body = f"Estado da visita actualizado para: {body.estado}."
                _asyncio2.create_task(_send_push(
                    user_id=str(analista.utilizador_id),
                    title=_push_title,
                    body=_push_body,
                    url=f"/visitas/{visita.id}",
                    db=db,
                ))
        except Exception:
            pass  # email/push failures must never break the state transition

    # In-app notifications (MensagemSistema) for coordinators/validators/clients
    _NOTIF_MAP: dict[str, tuple[tuple[str, ...], str, str]] = {}
    if body.estado == "inserida":
        _NOTIF_MAP = {
            "inserida": (
                ("coordenador", "validador"),
                f"Visita #{visita.id} submetida",
                f"A visita #{visita.id} do estudo foi submetida e aguarda validação.",
            )
        }
    elif body.estado in ("corrigir", "corrigir_email"):
        _label = "correcção por email" if body.estado == "corrigir_email" else "correcção"
        _NOTIF_MAP = {
            body.estado: (
                ("coordenador",),
                f"Visita #{visita.id} devolvida para {_label}",
                f"A visita #{visita.id} foi devolvida ao analista para {_label}.",
            )
        }
    elif body.estado == "validada":
        _NOTIF_MAP = {
            "validada": (
                ("coordenador",),
                f"Visita #{visita.id} validada",
                f"A visita #{visita.id} foi validada com sucesso.",
            )
        }
    elif body.estado == "fechada":
        _NOTIF_MAP = {
            "fechada": (
                ("coordenador", "cliente"),
                f"Visita #{visita.id} concluída",
                f"A visita #{visita.id} foi encerrada e os resultados estão disponíveis.",
            )
        }

    if _NOTIF_MAP:
        try:
            from app.models.message import MensagemSistema
            from app.models.user import PermissaoEstudo
            import asyncio as _asyncio_push
            from app.services.push_service import send_push

            roles, _assunto, _corpo = _NOTIF_MAP[body.estado]
            _perms_q = await db.execute(
                select(PermissaoEstudo).where(
                    PermissaoEstudo.estudo_id == visita.estudo_id,
                    PermissaoEstudo.role.in_(roles),
                )
            )
            for _perm in _perms_q.scalars().all():
                if _perm.utilizador_id != user.id:  # don't notify the actor
                    db.add(MensagemSistema(
                        remetente_id=user.id,
                        destinatario_id=_perm.utilizador_id,
                        assunto=_assunto,
                        corpo=_corpo,
                    ))
                    # Web Push for each notified user
                    _asyncio_push.create_task(send_push(
                        user_id=str(_perm.utilizador_id),
                        title=_assunto,
                        body=_corpo[:120],
                        url=f"/visitas/{visita.id}",
                        db=db,
                    ))
        except Exception:
            pass  # notification failures must never break the state transition

    # Cognira IA — auto-run validation when visit is submitted (inserida)
    # Fires in background; result stored via save_to_db=True on the visit row
    if body.estado == "inserida":
        import asyncio as _asyncio
        _vid = visita.id

        async def _bg_ia_validate() -> None:
            try:
                from app.database import async_session as _async_session
                from app.ai.intelligence import validar_visita_assistido as _vai
                async with _async_session() as bg_db:
                    await _vai(visita_id=_vid, db=bg_db, save_to_db=True)
            except Exception:
                pass  # IA failures must never break the state transition

        _asyncio.create_task(_bg_ia_validate())

    # Real-time: notify study participants about the state change
    try:
        from app.ws import manager
        from app.models.user import PermissaoEstudo
        _ws_roles = ("coordenador", "validador", "cliente", "analista")
        _perms_ws_q = await db.execute(
            select(PermissaoEstudo.utilizador_id).where(
                PermissaoEstudo.estudo_id == visita.estudo_id,
                PermissaoEstudo.role.in_(_ws_roles),
            )
        )
        _ws_ids = [str(uid) for uid in _perms_ws_q.scalars().all() if uid != user.id]
        await manager.broadcast_to_users(
            _ws_ids,
            {"evento": "visita_estado", "visita_id": visita.id, "estado": body.estado},
        )
    except Exception:
        pass

    return visita


# ── Export ────────────────────────────────────────────────────────────────────

@router.get("/export/excel")
async def export_visitas_excel(
    estudo_id: Optional[int] = None,
    onda_id: Optional[int] = None,
    estado: Optional[str] = None,
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export visitas to Excel (.xlsx). Filter by estudo, onda and/or estado."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    q = select(Visita)
    _tid_exp = tenant_filter(user)
    if _tid_exp is not None:
        from app.models.study import Estudo as _EstudoX
        from app.models.client import Cliente as _ClienteX
        q = q.join(_EstudoX, Visita.estudo_id == _EstudoX.id).join(_ClienteX, _EstudoX.cliente_id == _ClienteX.id).where(_ClienteX.tenant_id == _tid_exp)
    if estudo_id:
        q = q.where(Visita.estudo_id == estudo_id)
    if onda_id:
        q = q.where(Visita.onda_id == onda_id)
    if estado:
        q = q.where(Visita.estado == estado)
    q = q.order_by(Visita.id)
    rows = (await db.execute(q)).scalars().all()

    # Pre-fetch analistas & estabelecimentos for all rows
    analista_map: dict[int, str] = {}
    estab_map: dict[int, str] = {}
    onda_map: dict[int, str] = {}

    analista_ids = {v.analista_id for v in rows if v.analista_id}
    estab_ids = {v.estabelecimento_id for v in rows}
    onda_ids = {v.onda_id for v in rows if v.onda_id}

    if analista_ids:
        from app.models.analyst import Analista
        res = await db.execute(select(Analista).where(Analista.id.in_(analista_ids)))
        for a in res.scalars().all():
            raw = a.nome
            analista_map[a.id] = pii.decrypt(raw) if isinstance(raw, (bytes, bytearray)) else str(raw or "")

    if estab_ids:
        from app.models.establishment import Estabelecimento
        res = await db.execute(select(Estabelecimento).where(Estabelecimento.id.in_(estab_ids)))
        for e in res.scalars().all():
            estab_map[e.id] = e.nome

    if onda_ids:
        res = await db.execute(select(Onda).where(Onda.id.in_(onda_ids)))
        for o in res.scalars().all():
            onda_map[o.id] = o.label

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Visitas"

    headers = [
        "ID", "Estudo ID", "Onda", "Estabelecimento", "Analista",
        "Estado", "Pontuação", "Tipo Visita", "Planeada Em",
        "Realizada Início", "Realizada Fim", "Inserida Em",
    ]
    header_fill = PatternFill("solid", fgColor="2D6BEE")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, v in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=v.id)
        ws.cell(row=row_idx, column=2, value=v.estudo_id)
        ws.cell(row=row_idx, column=3, value=onda_map.get(v.onda_id, "") if v.onda_id else "")
        ws.cell(row=row_idx, column=4, value=estab_map.get(v.estabelecimento_id, ""))
        ws.cell(row=row_idx, column=5, value=analista_map.get(v.analista_id, "") if v.analista_id else "")
        ws.cell(row=row_idx, column=6, value=v.estado)
        ws.cell(row=row_idx, column=7, value=float(v.pontuacao) if v.pontuacao else None)
        ws.cell(row=row_idx, column=8, value=v.tipo_visita)
        ws.cell(row=row_idx, column=9, value=v.planeada_em.isoformat() if v.planeada_em else None)
        ws.cell(row=row_idx, column=10, value=v.realizada_inicio.isoformat() if v.realizada_inicio else None)
        ws.cell(row=row_idx, column=11, value=v.realizada_fim.isoformat() if v.realizada_fim else None)
        ws.cell(row=row_idx, column=12, value=v.inserida_em.isoformat() if v.inserida_em else None)

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"visitas_estudo{estudo_id or 'todos'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── PDF export individual ─────────────────────────────────────────────────────

@router.get("/{visita_id}/pdf")
async def export_visita_pdf(
    visita_id: int,
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Generate a single-visit PDF report including scores, responses and photos."""
    # ── Fetch visita ─────────────────────────────────────────────────────────
    visita = (await db.execute(select(Visita).where(Visita.id == visita_id))).scalar_one_or_none()
    if not visita:
        raise HTTPException(404, "Visita não encontrada")
    # Tenant isolation
    _tid_pdf = tenant_filter(user)
    if _tid_pdf is not None:
        from app.models.study import Estudo as _EstudoPDF
        from app.models.client import Cliente as _ClientePDF
        _own_pdf = (await db.execute(
            select(_ClientePDF.id)
            .join(_EstudoPDF, _EstudoPDF.cliente_id == _ClientePDF.id)
            .where(_EstudoPDF.id == visita.estudo_id, _ClientePDF.tenant_id == _tid_pdf)
        )).scalar_one_or_none()
        if _own_pdf is None:
            raise HTTPException(status_code=403, detail="Acesso negado")

    analista_nome = ""
    if visita.analista_id:
        a = (await db.execute(select(Analista).where(Analista.id == visita.analista_id))).scalar_one_or_none()
        if a:
            raw = a.nome
            analista_nome = pii.decrypt(raw) if isinstance(raw, (bytes, bytearray)) else str(raw or "")

    estab_nome = ""
    e = (await db.execute(select(Estabelecimento).where(Estabelecimento.id == visita.estabelecimento_id))).scalar_one_or_none()
    if e:
        estab_nome = e.nome or ""

    onda_label = ""
    if visita.onda_id:
        o = (await db.execute(select(Onda).where(Onda.id == visita.onda_id))).scalar_one_or_none()
        if o:
            onda_label = o.label or ""

    cc = (await db.execute(
        select(CaracterizacaoCache).where(CaracterizacaoCache.visita_id == visita_id)
    )).scalar_one_or_none()
    caracterizacao: dict = cc.dados if cc and cc.dados else {}

    from app.models.photo import FotoVisita
    fotos = (await db.execute(
        select(FotoVisita).where(FotoVisita.visita_id == visita_id)
    )).scalars().all()

    # ── Build PDF ─────────────────────────────────────────────────────────────
    from fpdf import FPDF
    from app.services import storage as _storage

    BUCKET = "fotos-visita"

    class RelatorioPDF(FPDF):
        def header(self):
            self.set_font("Helvetica", "B", 12)
            self.set_fill_color(30, 80, 160)
            self.set_text_color(255, 255, 255)
            self.cell(0, 10, "  Relatório de Visita", fill=True, ln=True)
            self.set_text_color(0, 0, 0)
            self.ln(2)

        def footer(self):
            self.set_y(-13)
            self.set_font("Helvetica", "I", 8)
            self.set_text_color(130, 130, 130)
            self.cell(0, 10, f"Pág. {self.page_no()}", align="C")

    pdf = RelatorioPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.set_margins(15, 15, 15)
    pdf.add_page()

    def section_title(title: str):
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_fill_color(230, 236, 248)
        pdf.cell(0, 7, f"  {title}", fill=True, ln=True)
        pdf.ln(1)

    def row(label: str, value: str):
        # Always reset X to left margin to avoid FPDFException from stray cursor positions
        if pdf.get_x() > pdf.l_margin + 0.5:
            pdf.set_x(pdf.l_margin)
        pdf.set_font("Helvetica", "B", 9)
        label_w = 55
        pdf.cell(label_w, 6, str(label)[:60], border=0)
        pdf.set_font("Helvetica", "", 9)
        # Use explicit remaining width (never pass 0 — multi_cell(0,...) can crash
        # if a prior drawing op left x at a non-margin position on the same line)
        value_w = pdf.w - pdf.l_margin - pdf.r_margin - label_w
        pdf.multi_cell(max(value_w, 20), 6, str(value or "—")[:1000], border=0)

    # ── Visit summary ─────────────────────────────────────────────────────────
    section_title("Dados da Visita")
    row("ID:", str(visita.id))
    row("Estado:", visita.estado)
    row("Tipo:", visita.tipo_visita)
    row("Estabelecimento:", estab_nome)
    row("Analista:", analista_nome)
    if onda_label:
        row("Onda:", onda_label)
    if visita.planeada_em:
        row("Planeada em:", visita.planeada_em.strftime("%d/%m/%Y %H:%M"))
    if visita.realizada_inicio:
        row("Realizada:", visita.realizada_inicio.strftime("%d/%m/%Y %H:%M"))
    if visita.pontuacao is not None:
        row("Pontuação:", f"{visita.pontuacao:.1f}")
    if visita.ia_veredicto:
        row("Veredicto IA:", visita.ia_veredicto)
    pdf.ln(4)

    # ── Caracterização ────────────────────────────────────────────────────────
    if caracterizacao:
        section_title("Respostas / Caracterização")
        for k, v in caracterizacao.items():
            row(f"{k}:", str(v))
        pdf.ln(4)

    # ── Photos ────────────────────────────────────────────────────────────────
    if fotos:
        section_title(f"Fotos ({len(fotos)})")
        pdf.ln(2)
        x_start = 15
        img_w = 55
        img_h = 42
        gap = 6
        col, max_cols = 0, 3

        for foto in fotos:
            try:
                img_bytes = _storage.download_bytes(BUCKET, foto.url_minio)
                import tempfile, os as _os
                suffix = ".jpg" if "jpeg" in (foto.mime_type or "") else ".png"
                with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                    tmp.write(img_bytes)
                    tmp_path = tmp.name
                x = x_start + col * (img_w + gap)
                if pdf.get_y() + img_h + 12 > pdf.h - 20:
                    pdf.add_page()
                    col = 0
                    x = x_start
                pdf.image(tmp_path, x=x, y=pdf.get_y(), w=img_w, h=img_h)
                _os.unlink(tmp_path)

                # IA badge under photo
                if foto.ia_veredicto:
                    badge_colors = {
                        "aprovada": (34, 197, 94),
                        "rejeitada": (239, 68, 68),
                        "inconclusiva": (234, 179, 8),
                    }
                    r, g, b = badge_colors.get(foto.ia_veredicto.lower(), (100, 100, 100))
                    pdf.set_xy(x, pdf.get_y() + img_h + 1)
                    pdf.set_font("Helvetica", "I", 7)
                    pdf.set_text_color(r, g, b)
                    pdf.cell(img_w, 4, foto.ia_veredicto, align="C")
                    pdf.set_text_color(0, 0, 0)
                    # Always reset X to left margin after badge so next row() call is safe
                    pdf.set_x(pdf.l_margin)

                col += 1
                if col >= max_cols:
                    col = 0
                    pdf.ln(img_h + 10)
            except Exception:
                col += 1
                if col >= max_cols:
                    col = 0
                    pdf.ln(img_h + 10)

        if col > 0:
            pdf.ln(img_h + 10)

    buf = BytesIO(pdf.output())
    filename = f"visita_{visita_id}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename={filename}"},
    )


# ---------------------------------------------------------------------------
# Cognira Intelligence — Module 6: Validation Assistant
# ---------------------------------------------------------------------------

from app.ai.intelligence import validar_visita_assistido, auto_qc_visita
from app.deps import require_role as _require_role


@router.post("/{visita_id}/validar-ia")
async def validar_visita_ia(
    visita_id: int,
    user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Cognira Module 6 — AI-assisted validation of a visit.

    Analyses responses and flags inconsistencies; returns recommendation (aprovar/corrigir/rever).
    """
    require_pro("ai_validation")
    _tid_vai = tenant_filter(user)
    if _tid_vai is not None:
        from app.models.study import Estudo as _EstudoVAI
        from app.models.client import Cliente as _ClienteVAI
        _v_vai = (await db.execute(select(Visita).where(Visita.id == visita_id))).scalar_one_or_none()
        if not _v_vai:
            raise HTTPException(404, "Visita não encontrada")
        _own_vai = (await db.execute(
            select(_ClienteVAI.id)
            .join(_EstudoVAI, _EstudoVAI.cliente_id == _ClienteVAI.id)
            .where(_EstudoVAI.id == _v_vai.estudo_id, _ClienteVAI.tenant_id == _tid_vai)
        )).scalar_one_or_none()
        if _own_vai is None:
            raise HTTPException(status_code=403, detail="Acesso negado")
    return await validar_visita_assistido(visita_id=visita_id, db=db)


# ── 8E.3 Auto-QC ─────────────────────────────────────────────────────────────

@router.post("/{visita_id}/auto-qc")
async def run_auto_qc(
    visita_id: int,
    user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Cognira Module 13 — AI auto-QC: flag suspicious or inconsistent responses."""
    require_pro("ai_validation")
    _tid_aqc = tenant_filter(user)
    if _tid_aqc is not None:
        from app.models.study import Estudo as _EstudoAQC
        from app.models.client import Cliente as _ClienteAQC
        _v_aqc = (await db.execute(select(Visita).where(Visita.id == visita_id))).scalar_one_or_none()
        if not _v_aqc:
            raise HTTPException(404, "Visita não encontrada")
        _own_aqc = (await db.execute(
            select(_ClienteAQC.id)
            .join(_EstudoAQC, _EstudoAQC.cliente_id == _ClienteAQC.id)
            .where(_EstudoAQC.id == _v_aqc.estudo_id, _ClienteAQC.tenant_id == _tid_aqc)
        )).scalar_one_or_none()
        if _own_aqc is None:
            raise HTTPException(status_code=403, detail="Acesso negado")
    return await auto_qc_visita(visita_id=visita_id, db=db)


# ── Wave 6.3 — GPS Proof-of-Presence ─────────────────────────────────────────

class _GpsCheckinBody(BaseModel):
    lat: float
    lon: float


@router.post("/{visita_id}/gps-checkin")
async def gps_checkin(
    visita_id: int,
    body: _GpsCheckinBody,
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Record GPS position when an analista checks in to start a visit.

    Returns the distance in metres between the check-in point and the
    establishment's registered coordinates (if available).
    """
    from datetime import timezone

    visita = (
        await db.execute(select(Visita).where(Visita.id == visita_id))
    ).scalar_one_or_none()
    if not visita:
        raise HTTPException(404, "Visita não encontrada")
    # Tenant isolation
    _tid_gps = tenant_filter(user)
    if _tid_gps is not None:
        from app.models.study import Estudo as _EstudoGPS
        from app.models.client import Cliente as _ClienteGPS
        _own_gps = (await db.execute(
            select(_ClienteGPS.id)
            .join(_EstudoGPS, _EstudoGPS.cliente_id == _ClienteGPS.id)
            .where(_EstudoGPS.id == visita.estudo_id, _ClienteGPS.tenant_id == _tid_gps)
        )).scalar_one_or_none()
        if _own_gps is None:
            raise HTTPException(status_code=403, detail="Acesso negado")

    visita.gps_checkin_lat = body.lat
    visita.gps_checkin_lon = body.lon
    visita.gps_checkin_em = datetime.now(timezone.utc)

    distancia_m: Optional[float] = None
    estab = (
        await db.execute(
            select(Estabelecimento).where(Estabelecimento.id == visita.estabelecimento_id)
        )
    ).scalar_one_or_none()
    if estab and estab.latitude and estab.longitude:
        distancia_m = _haversine_m(
            (body.lat, body.lon),
            (float(estab.latitude), float(estab.longitude)),
        )

    await db.commit()
    return {
        "visita_id": visita_id,
        "gps_checkin_lat": body.lat,
        "gps_checkin_lon": body.lon,
        "distancia_m": distancia_m,
    }


def _haversine_m(a: tuple[float, float], b: tuple[float, float]) -> float:
    """Return great-circle distance in metres between two (lat, lon) pairs."""
    R = 6_371_000
    lat1, lon1 = map(radians, a)
    lat2, lon2 = map(radians, b)
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    h = sin(dlat / 2) ** 2 + cos(lat1) * cos(lat2) * sin(dlon / 2) ** 2
    return 2 * R * atan2(sqrt(h), sqrt(1 - h))

