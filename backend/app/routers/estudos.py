from io import BytesIO
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select, func, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.deps import get_current_user, require_role, tenant_filter
from app.models.study import Estudo, Onda, FiltroEstudo
from app.models.client import Cliente
from app.models.user import Utilizador, PermissaoEstudo
from app.models.evaluation import Grelha, SecaoGrelha, CriterioGrelha
from app.schemas import EstudoCreate, EstudoOut, OndaOut, GrelhaCreate, GrelhaOut
from app.ai.intelligence import (
    gerar_relatorio_narrativo, gerar_insights, planear_visitas_automatico,
    gerar_word_cloud, comparativo_temporal, analisar_sentimento,
)

from app.models.visit import Visita
from app.edition import require_pro
from app.services.estudos_service import estudo_or_404 as _estudo_or_404, compute_benchmarking

router = APIRouter()


@router.get("/", response_model=list[EstudoOut])
async def list_estudos(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    is_staff = user.role_global in ("admin", "coordenador")
    q = select(Estudo).order_by(Estudo.id)
    tid = tenant_filter(user)
    if not is_staff:
        # Restrict to estudos where the user has an explicit permission entry
        allowed_ids = [p.estudo_id for p in (user.permissoes or [])]
        if not allowed_ids:
            return []
        q = q.where(Estudo.id.in_(allowed_ids))
    elif tid is not None:
        # Staff admin/coordenador: scope to their tenant's clients
        q = q.join(Cliente, Estudo.cliente_id == Cliente.id).where(Cliente.tenant_id == tid)
    q = q.offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(q)
    estudos = result.scalars().all()

    # Batch visita counts
    estudo_ids = [e.id for e in estudos]
    visita_counts: dict[int, int] = {}
    if estudo_ids:
        vc_rows = await db.execute(
            select(Visita.estudo_id, func.count(Visita.id).label("cnt"))
            .where(Visita.estudo_id.in_(estudo_ids))
            .group_by(Visita.estudo_id)
        )
        visita_counts = {row.estudo_id: row.cnt for row in vc_rows.all()}

    items = []
    for e in estudos:
        from app.schemas import EstudoOut
        out = EstudoOut.model_validate(e)
        out.total_visitas = visita_counts.get(e.id, 0)
        items.append(out)
    return items


# ---------------------------------------------------------------------------
# Benchmarking cross-estudo (anonymised)  (MUST be before /{estudo_id})
# ---------------------------------------------------------------------------

@router.get("/benchmarking")
async def benchmarking_estudos(
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await compute_benchmarking(db, user)


@router.get("/{estudo_id}", response_model=EstudoOut)
async def get_estudo(
    estudo_id: int,
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    estudo = await _estudo_or_404(db, estudo_id, user)
    # Non-staff users may only access studies they have an explicit permission for
    if user.role_global not in ("admin", "coordenador"):
        allowed_ids = [p.estudo_id for p in (user.permissoes or [])]
        if estudo_id not in allowed_ids:
            raise HTTPException(status_code=403, detail="Sem acesso a este estudo")
    return estudo


@router.post("/", response_model=EstudoOut, status_code=201)
async def create_estudo(
    body: EstudoCreate,
    user: Utilizador = Depends(require_role("admin", "coordenador")),
    db: AsyncSession = Depends(get_db),
):
    # Verify the chosen cliente belongs to the user's tenant
    tid = tenant_filter(user)
    if tid is not None:
        cq = select(Cliente).where(Cliente.id == body.cliente_id, Cliente.tenant_id == tid)
        if not (await db.execute(cq)).scalar_one_or_none():
            raise HTTPException(status_code=404, detail="Cliente não encontrado")
    estudo = Estudo(**body.model_dump())
    db.add(estudo)
    await db.flush()
    await db.refresh(estudo)
    return estudo


@router.put("/{estudo_id}", response_model=EstudoOut)
async def update_estudo(
    estudo_id: int,
    body: EstudoCreate,
    user: Utilizador = Depends(require_role("admin", "coordenador")),
    db: AsyncSession = Depends(get_db),
):
    estudo = await _estudo_or_404(db, estudo_id, user)
    for k, v in body.model_dump().items():
        setattr(estudo, k, v)
    await db.flush()
    await db.refresh(estudo)
    return estudo


@router.delete("/{estudo_id}", status_code=204)
async def delete_estudo(
    estudo_id: int,
    user: Utilizador = Depends(require_role("admin", "coordenador")),
    db: AsyncSession = Depends(get_db),
):
    estudo = await _estudo_or_404(db, estudo_id, user)
    await db.delete(estudo)
    await db.commit()


@router.get("/{estudo_id}/ondas", response_model=list[OndaOut])
async def list_ondas(
    estudo_id: int,
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Onda).where(Onda.estudo_id == estudo_id).order_by(Onda.id)
    )
    return result.scalars().all()


# ── PDF report ────────────────────────────────────────────────────────────────

_PDF_LABELS: dict[str, dict[str, str]] = {
    "pt": {
        "title": "Relatorio do Estudo",
        "nome": "Nome",
        "descricao": "Descricao",
        "estado": "Estado",
        "resumo": "Resumo de Visitas",
        "total": "Total visitas",
        "media": "Pontuacao media",
        "col_estado": "Estado",
        "col_visitas": "Visitas",
        "ondas": "Ondas",
        "gerado": "Gerado em",
    },
    "en": {
        "title": "Study Report",
        "nome": "Name",
        "descricao": "Description",
        "estado": "Status",
        "resumo": "Visit Summary",
        "total": "Total visits",
        "media": "Average score",
        "col_estado": "Status",
        "col_visitas": "Visits",
        "ondas": "Waves",
        "gerado": "Generated on",
    },
    "es": {
        "title": "Informe del Estudio",
        "nome": "Nombre",
        "descricao": "Descripcion",
        "estado": "Estado",
        "resumo": "Resumen de Visitas",
        "total": "Total visitas",
        "media": "Puntuacion media",
        "col_estado": "Estado",
        "col_visitas": "Visitas",
        "ondas": "Oleadas",
        "gerado": "Generado el",
    },
    "fr": {
        "title": "Rapport de l'Etude",
        "nome": "Nom",
        "descricao": "Description",
        "estado": "Etat",
        "resumo": "Resume des Visites",
        "total": "Total visites",
        "media": "Score moyen",
        "col_estado": "Etat",
        "col_visitas": "Visites",
        "ondas": "Vagues",
        "gerado": "Genere le",
    },
}


@router.get("/{estudo_id}/relatorio/pdf")
async def export_estudo_pdf(
    estudo_id: int,
    locale: str = Query("pt"),
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Generate a PDF summary report for a study."""
    try:
        from fpdf import FPDF
    except ImportError:
        raise HTTPException(status_code=500, detail="fpdf2 not installed")

    lbl = _PDF_LABELS.get(locale, _PDF_LABELS["pt"])

    estudo = await _estudo_or_404(db, estudo_id, user)

    # Fetch stats
    from app.models.visit import Visita
    q_states = select(Visita.estado, func.count(Visita.id).label("cnt")).where(
        Visita.estudo_id == estudo_id
    ).group_by(Visita.estado)
    state_rows = (await db.execute(q_states)).all()
    por_estado = {r.estado: r.cnt for r in state_rows}
    total = sum(por_estado.values())

    q_avg = select(func.avg(Visita.pontuacao)).where(Visita.estudo_id == estudo_id)
    avg_val = (await db.execute(q_avg)).scalar_one_or_none()
    pontuacao_media = round(float(avg_val), 1) if avg_val else None

    ondas = (await db.execute(
        select(Onda).where(Onda.estudo_id == estudo_id).order_by(Onda.id)
    )).scalars().all()

    # Build PDF
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 20)
    pdf.set_text_color(255, 51, 0)
    pdf.cell(0, 12, f"{lbl['title']} #{estudo_id}", ln=True)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 12)
    pdf.cell(0, 8, f"{lbl['nome']}: {estudo.nome}", ln=True)
    if getattr(estudo, 'descricao', None):
        pdf.cell(0, 8, f"{lbl['descricao']}: {estudo.descricao[:120]}", ln=True)
    pdf.cell(0, 8, f"{lbl['estado']}: {estudo.estado}", ln=True)
    pdf.ln(4)

    # Stats section
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, lbl["resumo"], ln=True)
    pdf.set_font("Helvetica", "", 12)
    pdf.cell(0, 8, f"{lbl['total']}: {total}", ln=True)
    if pontuacao_media is not None:
        pdf.cell(0, 8, f"{lbl['media']}: {pontuacao_media}", ln=True)
    pdf.ln(2)

    # States table
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_fill_color(255, 51, 0)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(90, 8, lbl["col_estado"], border=1, fill=True)
    pdf.cell(40, 8, lbl["col_visitas"], border=1, fill=True, ln=True)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 11)
    for estado, cnt in sorted(por_estado.items(), key=lambda x: -x[1]):
        pct = round(cnt / total * 100, 1) if total else 0
        pdf.cell(90, 7, f"  {estado}", border=1)
        pdf.cell(40, 7, f"{cnt}  ({pct}%)", border=1, ln=True)
    pdf.ln(6)

    # Ondas
    if ondas:
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, lbl["ondas"], ln=True)
        pdf.set_font("Helvetica", "", 11)
        for o in ondas:
            pdf.cell(0, 7, f"  - {o.label}", ln=True)

    buf = BytesIO(pdf.output())
    filename = f"relatorio_estudo_{estudo_id}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Campo configuration ───────────────────────────────────────────────────────

class CampoConfig(BaseModel):
    chave: str
    label: str
    tipo: str = "text"          # text | number | select | boolean
    opcoes: List[str] = []      # for tipo=select
    obrigatorio: bool = False


def _parse_campos(tipo_car: Optional[dict]) -> list[dict]:
    """Convert any stored format to List[CampoConfig-dict]."""
    if not tipo_car:
        return []
    if "v2" in tipo_car:
        return tipo_car["v2"]
    # Old format: {"0": "label", "1": "label", ...}
    return [
        {"chave": label.lower().replace(" ", "_").replace("(", "").replace(")", ""),
         "label": label, "tipo": "text", "opcoes": [], "obrigatorio": False}
        for _, label in sorted(tipo_car.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 9999)
    ]


def _has_study_access(user: Utilizador, estudo_id: int, db_sync_permissoes) -> bool:
    """Check if user has any permission on this study (used inside async context)."""
    return any(p.estudo_id == estudo_id for p in (db_sync_permissoes or []))


@router.get("/{estudo_id}/campos", response_model=List[CampoConfig])
async def get_campos_configuracao(
    estudo_id: int,
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Return the field configuration for visits in this study."""
    estudo = await _estudo_or_404(db, estudo_id, user)
    return _parse_campos(estudo.tipo_caracterizacao)


@router.put("/{estudo_id}/campos", response_model=List[CampoConfig])
async def put_campos_configuracao(
    estudo_id: int,
    campos: List[CampoConfig],
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Replace the field configuration for visits in this study.
    Admin/coordenador can always edit. Cliente can edit only their own studies.
    """
    estudo = await _estudo_or_404(db, estudo_id, user)

    # Access control
    is_staff = user.role_global in ("admin", "coordenador")
    if not is_staff:
        perm = (await db.execute(
            select(PermissaoEstudo).where(
                PermissaoEstudo.utilizador_id == user.id,
                PermissaoEstudo.estudo_id == estudo_id,
            )
        )).scalar_one_or_none()
        if not perm:
            raise HTTPException(status_code=403, detail="Sem acesso a este estudo")

    # Validate chaves are unique
    chaves = [c.chave for c in campos]
    if len(chaves) != len(set(chaves)):
        raise HTTPException(status_code=400, detail="Chaves duplicadas nos campos")

    # Save as v2 format
    estudo.tipo_caracterizacao = {"v2": [c.model_dump() for c in campos]}
    await db.flush()
    await db.refresh(estudo)
    return _parse_campos(estudo.tipo_caracterizacao)


# ---------------------------------------------------------------------------
# Cognira Intelligence — Module 1: Relatório Narrativo por Onda
# ---------------------------------------------------------------------------

@router.post("/{estudo_id}/relatorio-ia")
async def gerar_relatorio_ia(
    estudo_id: int,
    onda_id: Optional[int] = Query(default=None, description="Filtrar por onda (omitir = estudo completo)"),
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Cognira Intelligence Module 1 — Generate a narrative AI report for a study/wave."""
    require_pro("ai_planning")
    estudo = await _estudo_or_404(db, estudo_id, user)

    is_staff = user.role_global in ("admin", "coordenador")
    if not is_staff:
        perm = (await db.execute(
            select(PermissaoEstudo).where(
                PermissaoEstudo.utilizador_id == user.id,
                PermissaoEstudo.estudo_id == estudo_id,
            )
        )).scalar_one_or_none()
        if not perm:
            raise HTTPException(status_code=403, detail="Sem acesso a este estudo")

    return await gerar_relatorio_narrativo(estudo_id, onda_id, db)


# ---------------------------------------------------------------------------
# Cognira Intelligence — Module 8: Planeamento Automático de Visitas
# ---------------------------------------------------------------------------

@router.post("/{estudo_id}/ondas/{onda_id}/planear-ia")
async def planear_visitas_ia(
    estudo_id: int,
    onda_id: int,
    user: Utilizador = Depends(require_role("admin", "coordenador")),
    db: AsyncSession = Depends(get_db),
):
    """Cognira Module 8 — AI-generated visit assignment plan for a specific onda."""
    require_pro("ai_planning")
    estudo = await _estudo_or_404(db, estudo_id, user)
    if not estudo:
        raise HTTPException(404, "Estudo não encontrado")
    return await planear_visitas_automatico(estudo_id=estudo_id, onda_id=onda_id, db=db)


class _PlanoEstab(BaseModel):
    id: int
    nome: str

class _PlanoItem(BaseModel):
    analista_id: int
    analista_nome: str
    estabelecimentos: list[_PlanoEstab]

class AplicarPlanoBody(BaseModel):
    plano: list[_PlanoItem]

@router.post("/{estudo_id}/ondas/{onda_id}/aplicar-plano", status_code=201)
async def aplicar_plano_ia(
    estudo_id: int,
    onda_id: int,
    body: AplicarPlanoBody,
    user: Utilizador = Depends(require_role("admin", "coordenador")),
    db: AsyncSession = Depends(get_db),
):
    """Apply an AI-generated plan: create planeada visits for each assignment."""
    estudo = await _estudo_or_404(db, estudo_id, user)
    if not estudo:
        raise HTTPException(404, "Estudo não encontrado")
    onda = (await db.execute(select(Onda).where(Onda.id == onda_id, Onda.estudo_id == estudo_id))).scalar_one_or_none()
    if not onda:
        raise HTTPException(404, "Onda não encontrada")

    criadas = 0
    ignoradas = 0
    for item in body.plano:
        for estab in item.estabelecimentos:
            # Skip if visit already exists for this onda+estabelecimento (not cancelled)
            existing = (await db.execute(
                select(Visita).where(
                    Visita.onda_id == onda_id,
                    Visita.estabelecimento_id == estab.id,
                    Visita.estado.notin_(["anulada", "cancelada"]),
                )
            )).scalar_one_or_none()
            if existing:
                ignoradas += 1
                continue
            v = Visita(
                estudo_id=estudo_id,
                onda_id=onda_id,
                analista_id=item.analista_id,
                estabelecimento_id=estab.id,
                estado="planeada",
            )
            db.add(v)
            criadas += 1

    await db.flush()
    return {"criadas": criadas, "ignoradas": ignoradas, "mensagem": f"{criadas} visita(s) criadas com estado 'planeada'. {ignoradas} ignorada(s) (já existiam)."}


# ---------------------------------------------------------------------------
# Cognira Intelligence — Module 5: Insights Semanais / On-Demand
# ---------------------------------------------------------------------------

@router.get("/{estudo_id}/insights")
async def get_insights_ia(
    estudo_id: int,
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Cognira Intelligence Module 5 — Generate real-time AI insights for a study."""
    require_pro("ai_planning")
    estudo = await _estudo_or_404(db, estudo_id, user)

    is_staff = user.role_global in ("admin", "coordenador")
    if not is_staff:
        perm = (await db.execute(
            select(PermissaoEstudo).where(
                PermissaoEstudo.utilizador_id == user.id,
                PermissaoEstudo.estudo_id == estudo_id,
            )
        )).scalar_one_or_none()
        if not perm:
            raise HTTPException(status_code=403, detail="Sem acesso a este estudo")

    return await gerar_insights(estudo_id, db)


# ── 8E.2 Word Cloud ───────────────────────────────────────────────────────────

@router.get("/{estudo_id}/word-cloud")
async def get_word_cloud(
    estudo_id: int,
    onda_id: Optional[int] = None,
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Extract keyword frequency cloud from open-text responses."""
    require_pro("ai_planning")
    await _check_estudo_access(estudo_id, user, db)
    return await gerar_word_cloud(estudo_id, db, onda_id)


# ── 8E.5 Comparativo Temporal ────────────────────────────────────────────────

@router.get("/{estudo_id}/comparativo-temporal")
async def get_comparativo_temporal(
    estudo_id: int,
    periodo_dias: int = Query(30, ge=7, le=180),
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """AI-generated narrative comparison between current and previous period."""
    await _check_estudo_access(estudo_id, user, db)
    return await comparativo_temporal(estudo_id, db, periodo_dias)


# ── 8E.1 Sentiment Analysis ──────────────────────────────────────────────────

@router.get("/{estudo_id}/sentimento")
async def get_sentimento(
    estudo_id: int,
    onda_id: Optional[int] = None,
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """NLP sentiment analysis over open-text visit responses."""
    require_pro("ai_planning")
    await _check_estudo_access(estudo_id, user, db)
    return await analisar_sentimento(estudo_id, db, onda_id)


# ── 8F.3 Scheduled / On-demand Report by Email ─────────────────────────────

class RelatorioEmailBody(BaseModel):
    destinatario: str
    onda_id: Optional[int] = None


@router.post("/{estudo_id}/relatorio-email")
async def enviar_relatorio_email(
    estudo_id: int,
    body: RelatorioEmailBody,
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """8F.3 — Generate narrative AI report and send to email with Excel attachment."""
    import smtplib
    from email.message import EmailMessage
    from sqlalchemy import select as _sel

    await _check_estudo_access(estudo_id, user, db)

    # Validate email target
    dest = body.destinatario.strip()
    if not dest or "@" not in dest:
        raise HTTPException(status_code=400, detail="Endereço de email inválido.")

    # Retrieve SMTP settings from configuracoes_sistema
    smtp_keys = ["smtp_host", "smtp_port", "smtp_user", "smtp_password", "smtp_from"]
    from app.models.settings import ConfiguracaoSistema
    rows = (await db.execute(
        select(ConfiguracaoSistema).where(ConfiguracaoSistema.chave.in_(smtp_keys))
    )).scalars().all()
    cfg = {r.chave: r.valor for r in rows}
    if not cfg.get("smtp_host") or not cfg.get("smtp_user") or not cfg.get("smtp_password"):
        raise HTTPException(
            status_code=422,
            detail="SMTP não configurado. Vai a Configurações → Sistema e define smtp_host, smtp_user e smtp_password.",
        )

    # Load study name
    estudo_row = await _estudo_or_404(db, estudo_id, user)
    if not estudo_row:
        raise HTTPException(status_code=404, detail="Estudo não encontrado.")
    estudo_nome = estudo_row.nome

    # Generate AI narrative report
    from app.ai.intelligence import gerar_relatorio_narrativo
    relatorio = await gerar_relatorio_narrativo(estudo_id, body.onda_id, db)
    texto_relatorio = relatorio.get("relatorio", "Relatório não disponível.")
    estatisticas = relatorio.get("estatisticas", {})

    # Build Excel attachment (reuse visit export logic)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl não instalado.")

    from app.models.visit import Visita
    q = select(Visita).where(Visita.estudo_id == estudo_id, Visita.activo == True)
    if body.onda_id:
        q = q.where(Visita.onda_id == body.onda_id)
    q = q.order_by(Visita.id)
    visitas = (await db.execute(q)).scalars().all()

    wb = openpyxl.Workbook()
    ws_vis = wb.active
    ws_vis.title = "Visitas"
    headers = ["ID", "Onda ID", "Estabelecimento ID", "Analista ID", "Estado", "Pontuação", "Tipo Visita", "Inserida Em"]
    hf = PatternFill("solid", fgColor="2D6BEE")
    hfont = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        c = ws_vis.cell(row=1, column=col, value=h)
        c.fill = hf; c.font = hfont; c.alignment = Alignment(horizontal="center")
    for ri, v in enumerate(visitas, 2):
        ws_vis.cell(ri, 1, v.id); ws_vis.cell(ri, 2, v.onda_id)
        ws_vis.cell(ri, 3, v.estabelecimento_id); ws_vis.cell(ri, 4, v.analista_id)
        ws_vis.cell(ri, 5, v.estado)
        ws_vis.cell(ri, 6, float(v.pontuacao) if v.pontuacao else None)
        ws_vis.cell(ri, 7, v.tipo_visita)
        ws_vis.cell(ri, 8, v.inserida_em.isoformat() if v.inserida_em else None)

    # Stats sheet
    ws_stats = wb.create_sheet("Estatísticas")
    ws_stats.append(["Indicador", "Valor"])
    for k, v_val in estatisticas.items():
        ws_stats.append([str(k), str(v_val)])

    buf = BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    # Build email
    msg = EmailMessage()
    msg["Subject"] = f"Relatório Cognira CX Intelligence — {estudo_nome}"
    msg["From"] = cfg.get("smtp_from") or cfg["smtp_user"]
    msg["To"] = dest
    html_body = f"""<html><body>
<h2 style="color:#2D6BEE">Cognira CX Intelligence — Relatório</h2>
<p><strong>Estudo:</strong> {estudo_nome}</p>
<hr>
<pre style="font-family:sans-serif;white-space:pre-wrap">{texto_relatorio}</pre>
<hr>
<p style="font-size:11px;color:#888">Gerado automaticamente pela plataforma Cognira CX Intelligence.<br>
Enviado a pedido de <em>{user.email or user.username}</em>.</p>
</body></html>"""
    msg.set_content(texto_relatorio)
    msg.add_alternative(html_body, subtype="html")
    filename = f"relatorio_{estudo_nome.replace(' ', '_')}.xlsx"
    msg.add_attachment(xlsx_bytes, maintype="application",
                       subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       filename=filename)

    # Send
    try:
        port = int(cfg.get("smtp_port") or 587)
        with smtplib.SMTP(cfg["smtp_host"], port, timeout=15) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.login(cfg["smtp_user"], cfg["smtp_password"])
            smtp.send_message(msg)
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Erro ao enviar email: {exc}")

    return {"status": "enviado", "destinatario": dest, "estudo": estudo_nome}


async def _check_estudo_access(estudo_id: int, user: Utilizador, db):
    """Raise 403/404 if user has no access to this study (includes tenant check)."""
    # Tenant isolation first — returns 404 (not 403) to avoid info disclosure
    await _estudo_or_404(db, estudo_id, user)
    is_staff = user.role_global in ("admin", "coordenador", "validador")
    if is_staff:
        return
    perm = (await db.execute(
        select(PermissaoEstudo).where(
            PermissaoEstudo.utilizador_id == user.id,
            PermissaoEstudo.estudo_id == estudo_id,
        )
    )).scalar_one_or_none()
    if not perm:
        raise HTTPException(status_code=403, detail="Sem acesso a este estudo")


# ── Grid (Grelha) CRUD ───────────────────────────────────────────────────────

@router.get("/{estudo_id}/grelhas", response_model=list[GrelhaOut])
async def list_grelhas(
    estudo_id: int,
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """List all evaluation grids for a study (with sections and criteria)."""
    await _check_estudo_access(estudo_id, user, db)
    rows = (await db.execute(
        select(Grelha).where(Grelha.estudo_id == estudo_id).order_by(Grelha.id)
    )).scalars().all()
    return rows


@router.post("/{estudo_id}/grelhas", response_model=GrelhaOut, status_code=201)
async def create_grelha(
    estudo_id: int,
    body: GrelhaCreate,
    user: Utilizador = Depends(require_role("admin", "coordenador")),
    db: AsyncSession = Depends(get_db),
):
    """Create a new evaluation grid for a study, including sections and criteria."""
    estudo = await _estudo_or_404(db, estudo_id, user)
    if not estudo:
        raise HTTPException(status_code=404, detail="Estudo não encontrado")

    grid = Grelha(
        estudo_id=estudo_id,
        nome=body.nome,
        versao=body.versao,
        tipo_visita=body.tipo_visita,
    )
    db.add(grid)
    await db.flush()

    for sec_data in body.secoes:
        secao = SecaoGrelha(
            grelha_id=grid.id,
            nome=sec_data.nome,
            ordem=sec_data.ordem,
            peso_secao=sec_data.peso_secao,
        )
        db.add(secao)
        await db.flush()
        for idx, crit_data in enumerate(sec_data.criterios):
            db.add(CriterioGrelha(
                grelha_id=grid.id,
                secao_id=secao.id,
                label=crit_data.label,
                peso=crit_data.peso,
                tipo=crit_data.tipo,
                ordem=crit_data.ordem if crit_data.ordem else idx,
            ))

    await db.flush()
    await db.refresh(grid)
    return grid


@router.get("/{estudo_id}/grelhas/{grelha_id}", response_model=GrelhaOut)
async def get_grelha(
    estudo_id: int,
    grelha_id: int,
    user: Utilizador = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await _check_estudo_access(estudo_id, user, db)
    grid = (await db.execute(
        select(Grelha).where(Grelha.id == grelha_id, Grelha.estudo_id == estudo_id)
    )).scalar_one_or_none()
    if not grid:
        raise HTTPException(status_code=404, detail="Grelha não encontrada")
    return grid


@router.put("/{estudo_id}/grelhas/{grelha_id}", response_model=GrelhaOut)
async def update_grelha(
    estudo_id: int,
    grelha_id: int,
    body: GrelhaCreate,
    user: Utilizador = Depends(require_role("admin", "coordenador")),
    db: AsyncSession = Depends(get_db),
):
    """Replace a grid's metadata, sections and criteria completely."""
    grid = (await db.execute(
        select(Grelha).where(Grelha.id == grelha_id, Grelha.estudo_id == estudo_id)
    )).scalar_one_or_none()
    if not grid:
        raise HTTPException(status_code=404, detail="Grelha não encontrada")

    grid.nome = body.nome
    grid.versao = body.versao
    grid.tipo_visita = body.tipo_visita

    # Remove existing sections (cascade deletes criteria)
    existing_secoes = (await db.execute(
        select(SecaoGrelha).where(SecaoGrelha.grelha_id == grelha_id)
    )).scalars().all()
    for s in existing_secoes:
        await db.delete(s)
    # Also delete orphan criteria (no secao_id)
    orphan_criteria = (await db.execute(
        select(CriterioGrelha).where(
            CriterioGrelha.grelha_id == grelha_id,
            CriterioGrelha.secao_id.is_(None),
        )
    )).scalars().all()
    for c in orphan_criteria:
        await db.delete(c)
    await db.flush()

    for sec_data in body.secoes:
        secao = SecaoGrelha(
            grelha_id=grid.id,
            nome=sec_data.nome,
            ordem=sec_data.ordem,
            peso_secao=sec_data.peso_secao,
        )
        db.add(secao)
        await db.flush()
        for idx, crit_data in enumerate(sec_data.criterios):
            db.add(CriterioGrelha(
                grelha_id=grid.id,
                secao_id=secao.id,
                label=crit_data.label,
                peso=crit_data.peso,
                tipo=crit_data.tipo,
                ordem=crit_data.ordem if crit_data.ordem else idx,
            ))

    await db.flush()
    await db.refresh(grid)
    return grid


@router.delete("/{estudo_id}/grelhas/{grelha_id}", status_code=204)
async def delete_grelha(
    estudo_id: int,
    grelha_id: int,
    user: Utilizador = Depends(require_role("admin", "coordenador")),
    db: AsyncSession = Depends(get_db),
):
    grid = (await db.execute(
        select(Grelha).where(Grelha.id == grelha_id, Grelha.estudo_id == estudo_id)
    )).scalar_one_or_none()
    if not grid:
        raise HTTPException(status_code=404, detail="Grelha não encontrada")
    await db.delete(grid)


